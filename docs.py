import streamlit as st
import pandas as pd
import json
from openpyxl import Workbook
import GenAI
import chardet
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
# newfeatures_comment="The dial pad functionality has been enhanced to support auto-complete for frequently dialed numbers. Additionally, there is a new feature that blocks calls to specific numbers after a certain number of failed attempts."

def testplan(key_suffix=""):
    uploaded_file = st.file_uploader(
        "Upload TestPlan",
        accept_multiple_files=True,
        type=["xlsx", "xls", "csv"],
        key=f"file_upload{key_suffix}"  # Unique key using suffix
    )
    
    if uploaded_file:
        if st.button("Submit Test Plan", key=f"submit_test_plan{key_suffix}"):
            st.session_state.uploaded_TestPlan = uploaded_file
            st.success("Test Plan submitted successfully")

    return st.session_state.get('uploaded_TestPlan', None)
def oldfeatures():
    # uploaded_file = st.file_uploader("Upload oldFeatures", accept_multiple_files=True, type=["txt"], key="file_upload_old_features")
    comment = st.text_area("Enter your details for old features", key="old_features_comments")
    
    if   comment:
        if st.button("Submit Old Features"):
            # st.session_state.uploaded_oldFeatures = uploaded_file
            st.session_state.oldFeature_comment = comment
            st.success("Old Features submitted successfully")
    
    return (st.session_state.get('uploaded_oldFeatures', None), st.session_state.get('oldFeature_comment', None))

def newfeatures():
    # uploaded_file = st.file_uploader("Upload newFeatures", accept_multiple_files=True, type=["txt"], key="file_upload_new_features")
    comment = st.text_area("Enter your details for new features", key="new_features_comments")
    
    if comment:
        if st.button("Submit New Features"):
            # st.session_state.uploaded_newFeatures = uploaded_file
            st.session_state.newFeature_comment = comment
            st.success("New Features submitted successfully")
    
    return (st.session_state.get('uploaded_newFeatures', None), st.session_state.get('newFeature_comment', None))


## method to convert csv,excel format to json
def convert_file_to_json(input_file, json_file):
    # Determine the file type and read the input file accordingly
    if input_file.type == 'text/csv':
        df = pd.read_csv(input_file)
    elif input_file.type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel']:
        df = pd.read_excel(input_file)
    else:
        st.error("Unsupported file type. Please upload a CSV or Excel file.")
        return
    # Initialize an empty list to hold the test cases
    test_cases = []
    current_case = None
    # Function to handle NaN values and return default empty strings
    def get_value(value):
        return value if pd.notna(value) else ""
    # Iterate over the rows in the DataFrame
    for _, row in df.iterrows():
        if pd.notna(row.get("ID")):  # Check if this row represents a new test case
            if current_case:
                # If there was a previous case, append it to the list
                test_cases.append(current_case)
            # Start a new test case
            current_case = {
                "testCaseID": str(int(get_value(row.get("ID")))),  # Convert ID to string, ensuring no float point
                "testCaseTitle": get_value(row.get("Title")),
                "application": get_value(row.get("Area Path")),  # Assuming 'Area Path' as 'application'
                "testType": get_value(row.get("Work Item Type")),  # Assuming 'Work Item Type' as 'testType'
                "responsibleTeams": [get_value(row.get("Assigned To"))],
                "testTags": [tag.strip() for tag in get_value(row.get("Tags")).split(';')] if pd.notna(row.get("Tags")) else [],
                "steps": []
            }
        if pd.notna(row.get("Test Step")):  # Check if this row contains a test step
            step = {
                "stepNumber": int(get_value(row.get("Test Step"))),
                "action": get_value(row.get("Step Action")),
                "expectedResult": [x.strip() for x in get_value(row.get("Step Expected")).split('\n')] if pd.notna(row.get("Step Expected")) else []
            }
            if current_case:  # Ensure that current_case is initialized before appending steps
                current_case["steps"].append(step)
    # Append the last test case if it exists
    if current_case:
        test_cases.append(current_case)
    # Write to JSON file
    with open(json_file, 'w') as f:
        json.dump(test_cases, f, indent=4)
    return json_file
"-------------------------------------------------------------------------------------------------------------------------------------------------"
# def process_and_convert_to_excel(newfeatures_comment,oldfeatures_comment,input_file_path='jsonfile.json', excel_file_path='formatted_test_cases.xlsx'):
#     """Process test cases from the input file, update them using the AI model, and save the results to an Excel file."""
#     try:
#         with open(input_file_path, 'r', encoding='utf-8', errors='ignore') as file:
#             all_test_cases = json.load(file)
        
#         updated_test_cases = []

#         if not all_test_cases:
#             st.warning("No test cases found in the input file.")
#             return
        
#         for test_case in all_test_cases:
#             single_test_case_json = json.dumps(test_case, indent=4)

            
#             try:
#                 response=GenAI.output_Generator(newfeatures_comment=newfeatures_comment,single_test_case=single_test_case_json)

#                 if isinstance(response, str) and response.strip():
#                     try:
#                         updated_test_case = json.loads(response)
#                         updated_test_cases.append(updated_test_case)
#                     except json.JSONDecodeError:
#                         print(f"Error decoding JSON response: {response}")
#                         continue
            
#             except Exception as e:
#                 st.write(f"Error processing test case: {e}")
#                 continue

#         if updated_test_cases:
#             flattened_data = []
#             for item in updated_test_cases:
#                 for step in item.get('steps', []):
#                     flattened_item = {
#                         'Test Case ID': item.get('testCaseID'),
#                         'Title': item.get('testCaseTitle'),
#                         'Application': item.get('application'),
#                         'Test Type': item.get('testType'),
#                         'Responsible Teams': ', '.join(item.get('responsibleTeams', [])),
#                         'Test Tags': ', '.join(item.get('testTags', [])),
#                         'Step Number': step.get('stepNumber'),
#                         'Action': step.get('action'),
#                         'Expected Result': ', '.join(step.get('expectedResult', []))  # Join list into a single string
#                     }
#                     flattened_data.append(flattened_item)
            
#             df = pd.DataFrame(flattened_data)

#             # Save DataFrame to Excel
#             wb = Workbook()
#             ws = wb.active
#             ws.title = 'Test Cases'

#             # Add header row
#             for col_num, column_title in enumerate(df.columns, 1):
#                 ws.cell(row=1, column=col_num, value=column_title)

#             # Add data rows
#             for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
#                 for col_num, value in enumerate(row_data, 1):
#                     ws.cell(row=row_num, column=col_num, value=value)

#             # Define border style
#             border_style = Border(
#                 left=Side(style='thin'),
#                 right=Side(style='thin'),
#                 top=Side(style='thin'),
#                 bottom=Side(style='thin')
#             )

#             # Function to merge cells for identical values
#             def merge_identical_cells(ws, column):
#                 previous_value = None
#                 start_row = 2  # Assuming header is in the first row
#                 merge_start = start_row

#                 for row in range(start_row, ws.max_row + 1):
#                     current_value = ws.cell(row=row, column=column).value
#                     if current_value != previous_value:
#                         if row - merge_start > 1:
#                             ws.merge_cells(start_row=merge_start, start_column=column, end_row=row - 1, end_column=column)
#                             # Center align and apply border to merged cells
#                             for merge_row in range(merge_start, row):
#                                 cell = ws.cell(row=merge_row, column=column)
#                                 cell.alignment = Alignment(horizontal='center', vertical='center')
#                                 cell.border = border_style
#                         merge_start = row
#                     previous_value = current_value

#                 # Handle the last range
#                 if ws.max_row - merge_start > 0:
#                     ws.merge_cells(start_row=merge_start, start_column=column, end_row=ws.max_row, end_column=column)
#                     for merge_row in range(merge_start, ws.max_row + 1):
#                         cell = ws.cell(row=merge_row, column=column)
#                         cell.alignment = Alignment(horizontal='center', vertical='center')
#                         cell.border = border_style

#             # Apply merging and formatting to all columns
#             for col in range(1, len(df.columns) + 1):
#                 merge_identical_cells(ws, col)

#             # Adjust column widths for better readability
#             for col in ws.columns:
#                 max_length = 0
#                 column = col[0].column_letter  # Get the column name
#                 for cell in col:
#                     try:
#                         if isinstance(cell.value, str):
#                             cleaned_value = cell.value.replace(';', '').strip()  # Remove semicolons and extra spaces
#                             cell.value = cleaned_value
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(cell.value)
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 ws.column_dimensions[column].width = adjusted_width

#             # Save the workbook
#             wb.save(excel_file_path)
#             st.info(f"Formatted and styled test cases have been saved to {excel_file_path}")

#         else:
#             st.warning("No valid test cases were processed.")

#     except (FileNotFoundError, json.JSONDecodeError) as e:
#         st.warning(f"Error processing file: {e}")
#     except Exception as e:
#         st.warning(f"An unexpected error occurred: {e}")

def process_and_convert_to_excel(newfeatures_comment, oldfeatures_comment, input_file_path='jsonfile.json', excel_file_path='formatted_test_cases.xlsx'):
    """Process test cases, update them using AI, and save to Excel."""
    test_cases_generator = load_test_cases(input_file_path)
    
    if not test_cases_generator:
        st.warning("No test cases found in the input file.")
        return
    
    updated_test_cases_generator = update_test_cases(test_cases_generator, newfeatures_comment)

    if updated_test_cases_generator:
        flattened_data_generator = flatten_test_cases(updated_test_cases_generator)
        format_and_save_to_excel(flattened_data_generator, excel_file_path)
    else:
        st.warning("No valid test cases were processed.")

def format_and_save_to_excel(flattened_data_generator, excel_file_path):
    """Format the flattened data and save it to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Cases'

    header = ['Test Case ID', 'Title', 'Application', 'Test Type', 'Responsible Teams', 'Test Tags', 'Step Number', 'Action', 'Expected Result']
    ws.append(header)

    # Define border style
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write rows iteratively
    for row_data in flattened_data_generator:
        row = [row_data[col] for col in header]
        ws.append(row)

    # Function to merge cells for identical values
    def merge_identical_cells(ws, column_index):
        """Merge cells in a column if they contain the same value."""
        previous_value = None
        merge_start = 2  # Assuming header is in the first row

        for row in range(2, ws.max_row + 1):
            current_value = ws.cell(row=row, column=column_index).value
            if current_value != previous_value:
                if row - merge_start > 1:
                    ws.merge_cells(start_row=merge_start, start_column=column_index, end_row=row - 1, end_column=column_index)
                    for r in range(merge_start, row):
                        ws.cell(row=r, column=column_index).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=r, column=column_index).border = border_style
                merge_start = row
            previous_value = current_value

        if ws.max_row - merge_start > 0:
            ws.merge_cells(start_row=merge_start, start_column=column_index, end_row=ws.max_row, end_column=column_index)
            for r in range(merge_start, ws.max_row + 1):
                ws.cell(row=r, column=column_index).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=r, column=column_index).border = border_style

    # Apply merging and formatting to all columns
    for col in range(1, len(header) + 1):
        merge_identical_cells(ws, col)

    # Adjust column widths for better readability
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_length

    # Save the workbook
    wb.save(excel_file_path)
    st.info(f"Formatted and styled test cases have been saved to {excel_file_path}")

def flatten_test_cases(updated_test_cases_generator):
    """Yield flattened test case data for Excel export."""
    for item in updated_test_cases_generator:
        test_case_id = item.get('testCaseID')
        title = item.get('testCaseTitle')
        application = item.get('application')
        test_type = item.get('testType')
        responsible_teams = ', '.join(item.get('responsibleTeams', []))
        test_tags = ', '.join(item.get('testTags', []))

        for step in item.get('steps', []):
            yield {
                'Test Case ID': test_case_id,
                'Title': title,
                'Application': application,
                'Test Type': test_type,
                'Responsible Teams': responsible_teams,
                'Test Tags': test_tags,
                'Step Number': step.get('stepNumber'),
                'Action': step.get('action'),
                'Expected Result': ', '.join(step.get('expectedResult', []))  # Join list into a single string
            }

def update_test_cases(test_cases_generator, newfeatures_comment):
    """Update test cases using the AI model, yielding each one to save memory."""
    for test_case in test_cases_generator:
        single_test_case_json = json.dumps(test_case)

        try:
            response = GenAI.output_Generator(newfeatures_comment=newfeatures_comment, single_test_case=single_test_case_json)

            if isinstance(response, str) and response.strip():
                try:
                    updated_test_case = json.loads(response)
                    yield updated_test_case
                except json.JSONDecodeError:
                    print(f"Error decoding JSON response: {response}")
                    continue
        
        except Exception as e:
            st.write(f"Error processing test case: {e}")
            continue


def load_test_cases(input_file_path):
    """Yield test cases from a JSON file one by one to save memory."""
    try:
        with open(input_file_path, 'r', encoding='utf-8', errors='ignore') as file:
            data = json.load(file)
            if isinstance(data, list):
                for test_case in data:
                    yield test_case
            else:
                st.warning("Invalid file format: Expected a list of test cases.")
                return
    except (FileNotFoundError, json.JSONDecodeError) as e:
        st.warning(f"Error loading test cases: {e}")

